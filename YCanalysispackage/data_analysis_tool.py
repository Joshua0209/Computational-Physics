'''
Author: YC (b10202042@g.ntu.edu.tw)
Used Packages: openpyxl, numpy, scipy, matplotlib, and uncertainties
'''

import openpyxl as xl
import numpy as np
import scipy.signal as spsgn
import scipy.optimize as spopt
import scipy.misc as spmisc
import scipy.stats as spstat
import matplotlib.pyplot as plt
import uncertainties as u
from uncertainties import unumpy as unp


## Auxiliary Functions ##

def _cells_to_array(cells):
    cells = np.array(cells)
    shape = np.shape(cells)
    cells = np.reshape(cells, (np.prod(shape),))
    for n in range(len(cells)):
        cells[n] = float((cells[n]).internal_value)
    cells = np.reshape(cells, shape)
    return cells

def _weight(arr, weight):
    shape = list(np.shape(arr))
    shape[0] -= len(weight)-1
    temp = np.zeros(shape)
    for n in range(len(weight)):
        temp += weight[n]*arr[n:len(arr)-len(weight)+1+n]
    return temp

def _self_neighbor(arr1d, radius):
    mark, record, num = {}, [0], 0
    for n in range(1, len(arr1d)):
        if arr1d[n]>(arr1d[record[0]]+radius/2.):
            mark[num] = np.array(record, dtype=int)
            num += 1
            record = [n]
        else: record += [n]
    mark[num] = np.array(record, dtype=int)

    bound = []
    for m in range(num+1):
        for ind in mark[m]:
            temp = []
            if (m-2) in mark: temp = (mark[m-2])[arr1d[mark[m-2]]>(arr1d[ind]-radius)]
            if not (len(temp)==0): lower = temp[0]
            elif (m-1) in mark: lower = (mark[m-1])[0]
            else: lower = (mark[m])[0]
            temp = []
            if (m+2) in mark: temp = (mark[m+2])[arr1d[mark[m+2]]<(arr1d[ind]+radius)]
            if not (len(temp)==0): upper = temp[-1]
            elif (m+1) in mark: upper = (mark[m+1])[-1]
            else: upper = (mark[m])[-1]
            bound += [[lower, upper]]
    
    return bound

def _neighbor(centers, arr1d, radius):
    mark, record, num, jump = {}, [0], 0, [arr1d[0]]
    for n in range(1, len(arr1d)):
        if arr1d[n]>(arr1d[record[0]]+radius/2.):
            mark[num] = np.array(record, dtype=int)
            num += 1
            record = [n]
            jump += [arr1d[n]]
        else: record += [n]
    mark[num] = np.array(record, dtype=int)

    ind = np.argsort(centers)
    centers = (np.array(centers))[ind]
    temp_bound, m = [], 0
    for center in centers:
        if m<num:
            while jump[m+1]<center:
                m += 1
                if m==num: break
        temp = []
        if m>=2: temp = (mark[m-2])[arr1d[mark[m-2]]>(center-radius)]
        if not (len(temp)==0): lower = temp[0]
        elif m>=1: lower = (mark[m-1])[0]
        else: lower = (mark[m])[0]
        temp = []
        if m<=num-2: temp = (mark[m+2])[arr1d[mark[m+2]]<(center+radius)]
        if not (len(temp)==0): upper = temp[-1]
        elif m<=num-1: upper = (mark[m+1])[-1]
        else: upper = (mark[m])[-1]
        temp_bound += [[lower, upper]]
    bound = np.zeros(np.shape(temp_bound), dtype=int)
    bound[ind] = temp_bound

    return bound

def _adjacent(centers, arr1d, window):
    sorted_ind = np.argsort(centers)
    centers = (np.array(centers))[sorted_ind]
    temp_starter = [0]
    for center in centers:
        ind, count = 0, 0
        for n in range(len(arr1d)):
            if count<window: count += 1
            else:
                if arr1d[n-window//2]<center: ind += 1
                else: break
        temp_starter += [temp_starter[-1]+ind]
        arr1d = arr1d[ind:]
    starter = np.zeros(len(centers), dtype=int)
    starter[sorted_ind] = temp_starter[1:]
    
    return starter

def _chisq(data, errsq, func, params=[]):
    return (((func(data.xval, *params)-data.yval)**2)/errsq).sum()


class Data:
    def __init__(self, xval, yval, xerr=np.nan, yerr=np.nan, scale=[1.,1.]):
        assert (np.shape(xval)==(len(yval),)) and (np.shape(yval)==(len(xval),)), "data error"
        self.xval = scale[0]*np.array(xval, dtype=float)
        self.yval = scale[1]*np.array(yval, dtype=float)

        if not (isinstance(xerr, list) or isinstance(xerr, np.ndarray)): self.xerr = np.full(len(xval), xerr, dtype=float)
        elif np.shape(xerr)==(len(xval),): self.xerr = scale[0]*np.array(xerr, dtype=float)
        else: assert False, "xerr error"

        if not (isinstance(yerr, list) or isinstance(yerr, np.ndarray)): self.yerr = np.full(len(yval), yerr, dtype=float)
        elif np.shape(yerr)==(len(yval),): self.yerr = scale[1]*np.array(yerr, dtype=float)
        else: assert False, "yerr error"

    def __add__(self, num):
        if isinstance(num, u.core.Variable) or isinstance(num, u.core.AffineScalarFunc): return Data(self.xval, (self.yval)+(num.n), self.xerr, np.sqrt((self.yerr)**2+(num.s)**2))
        elif isinstance(num, Data):
            self.checkequalx(num)
            return Data(self.xval, (self.yval)+(num.yval), self.xerr, np.sqrt((self.yerr)**2+(num.yerr)**2))
        else: return Data(self.xval, (self.yval)+num, self.xerr, self.yerr)

    def __radd__(self, num):
        return self+num

    def __sub__(self, num):
        return self+(num*(-1.))

    def __rsub__(self, num):
        return (-self)+num

    def __neg__(self):
        return Data(self.xval, -self.yval, self.xerr, self.yerr)

    def __mul__(self, num):
        if isinstance(num, u.core.Variable) or isinstance(num, u.core.AffineScalarFunc): return Data(self.xval, (num.n)*(self.yval), self.xerr, np.sqrt((num.n)**2*(self.yerr)**2+(num.s)**2*(self.yval)**2))
        elif isinstance(num, Data):
            self.checkequalx(num)
            return Data(self.xval, (self.yval)*(num.yval), self.xerr, np.sqrt(((self.yval)*(num.yerr))**2+((self.yerr)*(num.yval))**2))
        else: return Data(self.xval, num*(self.yval), self.xerr, abs(num)*(self.yerr))

    def __rmul__(self, num):
        return self*num

    def __truediv__(self, num):
        if isinstance(num, Data):
            self.checkequalx(num)
            return Data(self.xval, (self.yval)/(num.yval), self.xerr, np.sqrt(((self.yerr)/(num.yval))**2+((num.yerr)*(self.yval)/(num.yval)**2)**2))
        else: return self*(1./num)

    def __rtruediv__(self, num):
        if isinstance(num, u.core.Variable) or isinstance(num, u.core.AffineScalarFunc): Data(self.xval, (num.n)/(self.yval), self.xerr, np.sqrt(((num.s)/(self.yval))**2+((self.yerr)*(num.n)/(self.yval)**2)**2))
        elif isinstance(num, Data):
            self.checkequalx(num)
            return Data(self.xval, (num.yval)/(self.yval), self.xerr, np.sqrt(((num.yerr)/(self.yval))**2+((self.yerr)*(num.yval)/(self.yval)**2)**2))
        else: return Data(self.xval, num/(self.yval), self.xerr, abs(num)*(self.yerr)/(self.yval)**2)

    def __pow__(self, num):
        assert num>=0., "exponent less than 0"
        assert isinstance(num, int) or isinstance(num, float) or isinstance(num, np.integer), "'num' error"
        if num==0.: return 1.
        else: return Data(self.xval, np.power(self.yval, num), self.xerr, num*(self.yerr)*np.abs(np.power(self.yval, num-1.)))

    def __gt__(self, other):
        return overwrite(self, other)

    def __rshift__(self, other):
        return xyzinterp(self, other)

    def __lshift__(self, other):
        return wxyinterp(other, self)

    def __abs__(self):
        return Data(self.xval, np.abs(self.yval), self.xerr, self.yerr)

    def __len__(self):
        return len(self.xval)

    def __str__(self):
        self.check()
        return str(self.uarray())

    def __getitem__(self, index):
        if isinstance(index, int) or isinstance(index, np.integer): return np.array([u.ufloat(self.xval[index], self.xerr[index]), u.ufloat(self.yval[index], self.yerr[index])])
        else: return Data(self.xval[index], self.yval[index], self.xerr[index], self.yerr[index])

    def __call__(self, num):
        if isinstance(num, u.core.Variable) or isinstance(num, u.core.AffineScalarFunc): return u.ufloat(*interpolate(self, num.n, num.s, extrap=True))
        else: return u.ufloat(*interpolate(self, num, extrap=True))

    def seterr(self, xerr=np.nan, yerr=np.nan):
        self.xerr = np.full(len(self), xerr)
        self.yerr = np.full(len(self), yerr)
    
    def setnan(self, xnum=np.nan, ynum=np.nan):
        self.xerr = np.nan_to_num(self.xerr, nan=xnum)
        self.yerr = np.nan_to_num(self.yerr, nan=ynum)

    def check(self):
        assert len(np.shape(self.xval))==1, "xval error"
        assert len(np.shape(self.yval))==1, "yval error"
        assert len(np.shape(self.xerr))==1, "xerr error"
        assert len(np.shape(self.yerr))==1, "yerr error"
        assert (len(self.xval)==len(self.yval)) and (len(self.xval)==len(self.xerr)) and (len(self.xval)==len(self.yerr)), "data error, length: {}".format([self.xval, self.yval, self.xerr, self.yerr])

    def checkequalx(self, other):
        assert (len(self)==len(other)), "size error"
        assert np.allclose(self.xval, other.xval, atol=0., equal_nan=True) and np.allclose(self.xerr, other.xerr, atol=0., equal_nan=True), "data error"

    def copy(self):
        return Data(np.array(self.xval, copy=True), np.array(self.yval, copy=True), np.array(self.xerr, copy=True), np.array(self.yerr, copy=True))

    def array(self, axis=None):
        return array(self, axis)

    def uarray(self, axis=None):
        return uarray(self, axis)

    def show(self, column=[1, 1, 0, 0]):
        self.check()
        temp = np.zeros((len(self), np.sum(column)))
        count = 0
        if column[0]:
            temp[:,count] = self.xval
            count += 1
        if column[1]:
            temp[:,count] = self.yval
            count += 1
        if column[2]:
            temp[:,count] = self.xerr
            count += 1
        if column[3]:
            temp[:,count] = self.yerr
            count += 1
        print(temp)
    
    def savetxt(self, file_name):
        savetxt(self, file_name)

    def savexlsx(self, file_name, sheet_name, start_cell):
        savexlsx(self, file_name, sheet_name, start_cell)

    def merge(self, other):
        self.xval = np.append(self.xval, other.xval)
        self.yval = np.append(self.yval, other.yval)
        self.xerr = np.append(self.xerr, other.xerr)
        self.yerr = np.append(self.yerr, other.yerr)

    def switch(self, to_sort=False):
        self.xval, self.yval, self.xerr, self.yerr = self.yval, self.xval, self.yerr, self.xerr
        if to_sort: self.sort()

    def amax(self, axis='y'):
        return amax(self, axis)
    
    def argmax(self, axis='y'):
        return argmax(self, axis)
    
    def amin(self, axis='y'):
        return amin(self, axis)
    
    def argmin(self, axis='y'):
        return argmin(self, axis)

    def sort(self, strict=True):
        temp = sort(self, strict)
        self.xval, self.yval, self.xerr, self.yerr = temp.xval, temp.yval, temp.xerr, temp.yerr

    def cut(self, itvls, axis='x', edge=True):
        temp = cut(self, itvls, axis, edge)
        self.xval, self.yval, self.xerr, self.yerr = temp.xval, temp.yval, temp.xerr, temp.yerr

    def snip(self, xitvl=[], yitvl=[]):
        temp = snip(self, xitvl, yitvl)
        self.xval, self.yval, self.xerr, self.yerr = temp.xval, temp.yval, temp.xerr, temp.yerr

    def apply(self, xfunc=None, yfunc=None):
        temp = apply(self, xfunc, yfunc)
        self.xval, self.yval, self.xerr, self.yerr = temp.xval, temp.yval, temp.xerr, temp.yerr

    def plot(self, xitvl=[], yitvl=[], graph=None, **kwargs):
        data = snip(self, xitvl, yitvl)

        if graph: return graph.plot(data.xval, data.yval, **kwargs)
        else: plt.plot(data.xval, data.yval, **kwargs)

    def scatter(self, xitvl=[], yitvl=[], graph=None, **kwargs):
        data = snip(self, xitvl, yitvl)

        if graph: return graph.scatter(data.xval, data.yval, **kwargs)
        else: plt.scatter(data.xval, data.yval, **kwargs)

    def errorbar(self, xitvl=[], yitvl=[], graph=None, **kwargs):
        data = snip(self, xitvl, yitvl)

        if graph: return graph.errorbar(data.xval, data.yval, xerr=data.xerr, yerr=data.yerr, **kwargs)
        else: plt.errorbar(data.xval, data.yval, xerr=data.xerr, yerr=data.yerr, **kwargs)

    def smooth(self, weight=[1., 1., 1., 1., 1.]):
        temp = smooth(self, weight)
        self.xval, self.yval, self.xerr, self.yerr = temp.xval, temp.yval, temp.xerr, temp.yerr

    def savgol(self, window, order):
        temp = savgol(self, window, order)
        self.xval, self.yval, self.xerr, self.yerr = temp.xval, temp.yval, temp.xerr, temp.yerr

    def nbrsmooth(self, radius):
        temp = nbrsmooth(self, radius)
        self.xval, self.yval, self.xerr, self.yerr = temp.xval, temp.yval, temp.xerr, temp.yerr

    def polysmooth(self, radius, order):
        temp = polysmooth(self, radius, order)
        self.xval, self.yval, self.xerr, self.yerr = temp.xval, temp.yval, temp.xerr, temp.yerr

    def fitting(self, func, threshold=1.5, **params):
        return fitting(self, func, threshold, **params)

    def polyfit(self, order):
        return polyfit(self, order)

    def linereg(self):
        return linereg(self)


class FitResult:
    def __init__(self, func, param, ave_chisq, rsq=0.):
        self.funcinfo = func
        if isinstance(func, str):
            if func=='line':
                temp = np.zeros((2,2))
                temp[0,0] = (param['m'])[0]
                temp[0,1] = (param['m'])[1]
                temp[1,0] = (param['b'])[0]
                temp[1,1] = (param['b'])[1]
                self.param = temp
                self.rsq = rsq
            elif func=='poly':
                order = len(param)-1
                temp = np.zeros((order+1,2))
                for n in range(order+1):
                    temp[order-n,0] = (param[n])[0]
                    temp[order-n,1] = (param[n])[1]
                self.param = temp
                self.rsq = -1.
            else: assert False, "'func' error"
        else:
            self.param = param
            self.rsq = -1.
        self.ave_chisq = ave_chisq
    
    def __int__(self):
        if isinstance(self.funcinfo, str):
            if self.funcinfo=='line': return 0
            else: return 1
        else: return 2
 
    def __str__(self):
        return str(self.makedict(True))

    def __getitem__(self, param):
        if param=='chisq/dof': return self.ave_chisq
        elif param=='rsq':
            assert int(self)==0, "'param' error"
            return self.rsq
        else:
            temp = self.makedict(True)
            assert param in temp, "'param' error"
            return temp[param]

    def __call__(self, num):
        return (self.func())(num)

    def makedict(self, aux=False):
        if int(self)==0:
            temp = {'m': u.ufloat(*((self.param)[0])), 'b': u.ufloat(*((self.param)[1]))}
            if aux:
                temp['chisq/dof'] = self.ave_chisq
                temp['rsq'] = self.rsq
        elif int(self)==1:
            temp = {}
            order = len(self.param)-1
            for n in range(order+1):
                temp['c'+str(order-n)] = u.ufloat(*((self.param)[n]))
            if aux: temp['chisq/dof'] = self.ave_chisq
        else:
            temp = {}
            for name in self.param:
                temp[name] = u.ufloat(*(self.param[name]))
            if aux: temp['chisq/dof'] = self.ave_chisq
        
        return temp

    def show(self):
        if int(self)==0: name = 'linear regression'
        elif int(self)==1: name = 'polynomial of order {}'.format(len(self.param)-1)
        else: name = (self.funcinfo).__name__
        temp_dict = self.makedict()

        print("####################")
        print("")
        print("Function:", name)
        print("")
        print("Parameters:")
        for key in temp_dict.keys():
            print(key+":", temp_dict[key])
        print("")
        print("Statistics:")
        print("chi-squared/dof:", self.ave_chisq)
        if int(self)==0: print("R-squared value:", self.rsq)
        print("")
        print("####################")

    def coeffs(self):
        if int(self)==2:
            temp = []
            for num in (self.param).values():
                temp += [num[0]]
            return np.array(temp)
        else: return (self.param)[:,0]

    def func(self, err=True):
        if err:
            if int(self)==2:
                func = lambda x: (self.funcinfo)(x, **(self.makedict()))
                return func
            else:
                coeffs = unp.uarray((self.param)[:,0], (self.param)[:,1])
                poly = np.poly1d(coeffs)
                return poly
        else:
            if int(self)==2:
                func = lambda x: (self.funcinfo)(x, *(self.val()))
                return func
            else: return np.poly1d((self.param)[:,0])

    def save(self, file_name):
        if int(self)==0: name = 'linear regression'
        elif int(self)==1: name = 'polynomial of order {}'.format(len(self.param)-1)
        else: name = (self.funcinfo).__name__
        temp_dict = self.makedict()
        
        if len(file_name)<4: file = open(file_name+'.txt', 'w')
        elif file_name[-4:]=='.txt': file = open(file_name, 'w')
        else: file = open(file_name+'.txt', 'w')
        file.write("Function: "+name+'\n')
        file.write("Parameters:"+'\n')
        for key in temp_dict.keys():
            file.write(key+": "+str((temp_dict[key]).n)+'+/-'+str((temp_dict[key]).s)+'\n')
        file.write("Statistics:"+'\n')
        file.write("chi-squared/dof: "+str(self.ave_chisq)+'\n')
        if int(self)==0: file.write("R-squared value: "+str(self.rsq)+'\n')
        file.write('\n')
        file.write(str(self.makedict(True)))
        file.close()

    def linspace(self, itvl, point=100):
        return linspace(self, itvl, point)


## Basic Manipulations ##

def array(data, axis=None):
    '''
    Return the ndarray of the Data object

    axis: 'x', 'y', and None
        whether return specific axis or not
    '''
    if axis=='x': return (np.dstack((data.xval, data.xerr)))[0]
    elif axis=='y': return (np.dstack((data.yval, data.yerr)))[0]
    else: return (np.dstack((data.xval, data.yval, data.xerr, data.yerr)))[0]

def uarray(data, axis=None):
    '''
    Return the uarray of the Data object

    axis: 'x', 'y', and None
        whether return specific axis or not
    '''
    if axis=='x': return unp.uarray(data.xval, data.xerr)
    elif axis=='y': return unp.uarray(data.yval, data.yerr)
    else:
        arr = array(data)
        return unp.uarray(arr[:,:2], arr[:,2:])

def makeData(*dataset, xerr=np.nan, yerr=np.nan, scale=[1., 1.]):
    '''
    Return a Data object from ndarray

    xerr: float
        error of the x-axis (only assigned if it is not given in dataset)

    yerr: float
        error of the y-axis (only assigned if it is not given in dataset)
    
    scale: 2 component list
        scale of the axes
    '''
    if len(dataset)==1:
        dataset = dataset[0]
        if (np.shape(dataset))[1]==2: return Data(dataset[:,0], dataset[:,1], xerr, yerr, scale)
        elif (np.shape(dataset))[1]==4: return Data(dataset[:,0], dataset[:,1], dataset[:,2], dataset[:,3], scale)
        else: assert False, "'dataset' error"
    elif len(dataset)==2:
        assert len(dataset[0])==len(dataset[1]), "'dataset' error"
        dataset1, dataset2 = dataset[0], dataset[1]
        if len(np.shape(dataset1))==1:
            x_val, x_err = dataset1, xerr
        elif len(np.shape(dataset1))==2:
            x_val, x_err = dataset1[:,0], dataset1[:,1]
        else: assert False, "'dataset' error"
        if len(np.shape(dataset2))==1:
            y_val, y_err = dataset2, yerr
        elif len(np.shape(dataset2))==2:
            y_val, y_err = dataset2[:,0], dataset2[:,1]
        else: assert False, "'dataset' error"
        
        return Data(x_val, x_err, y_val, y_err, scale)
    else: assert False, "'dataset' error"

def umakeData(*dataset, scale=[1., 1.]):
    '''
    Return a Data object from uarray
    
    scale: 2 component list
        scale of the axes
    '''
    if len(dataset)==1:
        val, err = unp.nominal_values(dataset[0]), unp.std_devs(dataset[0])
        return Data(val[:,0], val[:,1], err[:,0], err[:,1], scale)
    elif len(dataset)==2:
        assert (len(dataset[0])==len(dataset[1])) and (len(np.shape(dataset[0]))==1) and (len(np.shape(dataset[1]))==1), "'dataset' error"
        return Data(unp.nominal_values(dataset[0]), unp.nominal_values(dataset[1]), unp.std_devs(dataset[0]), unp.std_devs(dataset[1]), scale)
    else: assert False, "'dataset' error"

def read(file_name, scale=[1., 1.], to_sort=False, **kwargs):
    '''
    Turn a file into a Data object

    scale: 2 component list
        scale of the axes

    to_sort: Boolean
        whether to sort the data or not

    kwargs: char, sheet_name, cells
        see 'readtxt' and 'readxlsx'
    '''
    assert len(file_name)>=4, "'file_name' error"
    if file_name[-4:]=='.txt':
        if 'char' in kwargs: temp = makeData(readtxt(file_name, kwargs['char']), scale=scale)
        else: temp = makeData(readtxt(file_name), scale=scale)
    else:
        assert len(file_name)>=5, "'file_name' error"
        if file_name[-5:]=='.xlsx': temp = makeData(readxlsx(file_name, kwargs['sheet_name'], kwargs['cells']), scale=scale)
        else: assert False, "'file_name' error"
    
    if to_sort: return sort(temp)
    else: return temp

def readtxt(file_name, char=[]):
    '''
    Turn a .txt file into a ndarray

    char: [char1, char2, ...]
        characters that split the context
    '''
    if len(file_name)<4:
        with open(file_name+'.txt') as file:
            lines = file.readlines()
    elif file_name[-4:]=='.txt':
        with open(file_name) as file:
            lines = file.readlines()
    else:
        with open(file_name+'.txt') as file:
            lines = file.readlines()
    temp = []
    for line in lines:
        temp += [line.split(*char)]
    temp = np.array(temp, dtype=float)

    return temp

def readxlsx(file_name, sheet_name, *cells):
    '''
    Turn a .xlsx file into a ndarray

    cells: str (e.g., "A1:B10" or "A1:A10", "C1:C10")
        focused cells in the work sheet
    '''
    if len(file_name)<5: wb = xl.load_workbook(file_name+'.xlsx')
    elif file_name[-5:]=='.xlsx': wb = xl.load_workbook(file_name)
    else: wb = xl.load_workbook(file_name+'.xlsx')
    ws = wb[sheet_name]
    record = []
    if isinstance(cells[0], list): cells = cells[0]
    for cell in cells:
        record += [_cells_to_array(ws[cell])]
    temp = record[0]
    if len(record)>1:
        for arr in record[1:]:
            temp = np.append(temp, arr, axis=1)
    wb.close()
    return temp

def readfit(file_name):
    '''
    Make a FitResult object from a .txt file
    '''
    if len(file_name)<4:
        with open(file_name+'.txt') as file:
            lines = file.readlines()
    elif file_name[-4:]=='.txt':
        with open(file_name) as file:
            lines = file.readlines()
    else:
        with open(file_name+'.txt') as file:
            lines = file.readlines()

def savetxt(data, file_name):
    '''
    Save a Data object as a .txt file
    '''
    if len(file_name)<4: np.savetxt(file_name+'.txt', array(data))
    elif file_name[-4:]=='.txt': np.savetxt(file_name, array(data))
    else: np.savetxt(file_name+'.txt', array(data))

def savexlsx(data, file_name, sheet_name, start_cell):
    '''
    Save a Data object as a .xlsx file

    start_cell: str (e.g., "C3")
        the location of the top-left cell
    '''
    try:
        if len(file_name)<5: wb = xl.load_workbook(file_name+'.xlsx')
        elif file_name[-5:]=='.xlsx': wb = xl.load_workbook(file_name)
        else: wb = xl.load_workbook(file_name+'.xlsx')
        try: ws = wb[sheet_name]
        except: ws = wb.create_sheet(sheet_name)
    except:
        wb = xl.Workbook()
        ws = wb.active
        ws.title = sheet_name
    
    r = (ws[start_cell]).row
    c = (ws[start_cell]).column
    arr = array(data)
    for m in range(len(data)):
        for n in range(4):
            ws.cell(r+m, c+n, arr[m,n])
    wb.save(file_name)
    wb.close()

def merge(*data):
    '''
    Merge several Data objects
    '''
    if len(data)==1: return data[0]
    else:
        new_data = data[0]
        for single_data in data[1:]:
            new_data.merge(single_data)

        return new_data

def switch(data, to_sort=False):
    '''
    Return a Data object with the x-axis and y-axis exchanged

    to_sort: Boolean
        whether to sort the return Data object or not
    '''
    if to_sort: return sort(Data(data.yval, data.xval, data.yerr, data.xerr))
    else: return Data(data.yval, data.xval, data.yerr, data.xerr)

def amax(data, axis='y'):
    '''
    Return the data point of the maximum

    axis: 'x' or 'y'
        the axis to be searched for maximum
    '''
    ind = argmax(data, axis)
    return data[ind]

def argmax(data, axis='y'):
    '''
    Return the index of the maximum

    axis: 'x' or 'y'
        the axis to be searched for maximum
    '''
    if axis=='x': return np.argmax(data.xval)
    elif axis=='y': return np.argmax(data.yval)
    else: assert False, "'axis' error"

def amin(data, axis='y'):
    '''
    Return the data point of the minimum

    axis: 'x' or 'y'
        the axis to be searched for minimum
    '''
    ind = argmin(data, axis)
    return data[ind]

def argmin(data, axis='y'):
    '''
    Return the index of the minimum

    axis: 'x' or 'y'
        the axis to be searched for minimum
    '''
    if axis=='x': return np.argmin(data.xval)
    elif axis=='y': return np.argmin(data.yval)
    else: assert False, "'axis' error"

def sort(data, strict=True):
    '''
    Return a Data object with the x-axis sorted

    strict: Boolean
        whether the x-axis is strictly increasing or not
    '''
    if strict:
        if np.amin(data.xval[1:]-data.xval[:-1])>0.: return data
        else:
            temp_xval, ind = np.unique(data.xval, return_inverse=True)
            temp_yval, temp_xerr, temp_yerr = np.zeros(len(temp_xval)), np.zeros(len(temp_xval)), np.zeros(len(temp_xval))
            for n in range(len(temp_xval)):
                yval, xerr, yerr = data.yval[np.isin(ind,n)], data.xerr[np.isin(ind,n)], data.yerr[np.isin(ind,n)]
                temp_yval[n] = np.average(yval)
                temp_xerr[n] = np.amax(xerr)
                temp_yerr[n] = np.sqrt((np.var(yval)+np.sum(yerr**2))/len(yerr))
            
            return Data(temp_xval, temp_yval, temp_xerr, temp_yerr)
    else:
        ind = np.argsort(data.xval)
        return Data(data.xval[ind], data.yval[ind], data.xerr[ind], data.yerr[ind])

def cut(data, itvls, axis='x', edge=True):
    '''
    Return a Data object with the range of axis limited

    itvls: a list of 2 component tuples (a,b)
        the restriction to the axis

    axis: 'x' or 'y'
        the restricted axis
    
    edge: Boolean
        whether to include the points on the edge or not
    '''
    assert isinstance(itvls, list), "'itvls' error"
    temp_xval, temp_yval, temp_xerr, temp_yerr = np.array([]), np.array([]), np.array([]), np.array([])
    for itvl in itvls:
        assert isinstance(itvl, tuple) and (len(itvl)==2), "'itvls' error"
        if axis=='x':
            if edge: indbool = (data.xval[:]-itvl[0])*(data.xval[:]-itvl[1])<=0.
            else: indbool = (data.xval[:]-itvl[0])*(data.xval[:]-itvl[1])<0.
        elif axis=='y':
            if edge: indbool = (data.yval[:]-itvl[0])*(data.yval[:]-itvl[1])<=0.
            else: indbool = (data.yval[:]-itvl[0])*(data.yval[:]-itvl[1])<0.
        else: assert False, "'axis' error"
        temp_xval = np.append(temp_xval, data.xval[indbool])
        temp_yval = np.append(temp_yval, data.yval[indbool])
        temp_xerr = np.append(temp_xerr, data.xerr[indbool])
        temp_yerr = np.append(temp_yerr, data.yerr[indbool])
    
    return Data(temp_xval, temp_yval, temp_xerr, temp_yerr)

def snip(data, xitvl=[], yitvl=[], edge=True):
    '''
    Return a Data object with points inside the XY square

    xitvl: 2 component list [a,b] or tuple (a,b)
        restriction to the x-axis (a < x < b)

    yitvl: 2 component list [a,b] or tuple (a,b)
        restriction to the y-axis (a < y < b)
    
    edge: Boolean
        whether to include the points on the edge or not
    '''
    if not (np.shape(xitvl)==(0,)):
        assert (np.shape(xitvl)==(2,)), "'xitvl' error"
        data = cut(data, [tuple(xitvl)], 'x', edge)
    if not (np.shape(yitvl)==(0,)):
        assert (np.shape(yitvl)==(2,)), "'yitvl' error"
        data = cut(data, [tuple(yitvl)], 'y', edge)
    
    return data

def overwrite(top_data, bottom_data):
    '''
    Return a Data object with top_data overwriting bottom_data in the overlapping region
    '''
    btm = [np.amin(bottom_data.xval), np.amax(bottom_data.xval)]
    tp = [np.amin(top_data.xval), np.amax(top_data.xval)]
    temp = Data([], [])
    if btm[0]<tp[0]: temp.merge(snip(bottom_data, [btm[0]-1., tp[0]], edge=False))
    temp.merge(top_data)
    if btm[1]>tp[1]: temp.merge(snip(bottom_data, [tp[1], btm[1]+1.], edge=False))

    return temp

def apply(data, xfunc=None, yfunc=None):
    '''
    Return a Data object after applying given functions on axes individually
    '''
    if xfunc==None:
        if yfunc==None: return data.copy()
        else:
            y_uarr = yfunc(uarray(data, 'y'))
            return Data(data.xval, unp.nominal_values(y_uarr), data.xerr, unp.std_devs(y_uarr))
    else:
        if yfunc==None:
            x_uarr = xfunc(uarray(data, 'x'))
            return Data(unp.nominal_values(x_uarr), data.yval, unp.std_devs(x_uarr), data.yerr)
        else:
            x_uarr, y_uarr = xfunc(uarray(data, 'x')), yfunc(uarray(data, 'y'))
            return Data(unp.nominal_values(x_uarr), unp.nominal_values(y_uarr), unp.std_devs(x_uarr), unp.std_devs(y_uarr))

def linspace(fitting_result, itvl, point=100):
    '''
    Return a Data object from a FitResult object in a given interval

    itvl: 2 component list [a,b] or tuple (a,b)
        the range of the x-axis

    point: int
        the number of points
    '''
    assert np.shape(itvl)==(2,), "'itvl' error"
    func = fitting_result.func()
    temp_x = np.linspace(itvl[0], itvl[1], point)
    temp_y = func(temp_x)

    return Data(temp_x, unp.nominal_values(temp_y), 0., unp.std_devs(temp_y))


## Plot Methods ##

def plot(data, xitvl=[], yitvl=[], graph=None, **kwargs):
    '''
    Return a Line2D object in Matplotlib or plot the Data object (graph = None)

    xitvl: 2 component list [a,b] or tuple (a,b)
        restriction to the x-axis (a < x < b)

    yitvl: 2 component list [a,b] or tuple (a,b)
        restriction to the y-axis (a < y < b)

    graph: a Axes object in Matplotlib
    '''
    data = snip(data, xitvl, yitvl)

    if graph: return graph.plot(data.xval, data.yval, **kwargs)
    else: plt.plot(data.xval, data.yval, **kwargs)

def scatter(data, xitvl=[], yitvl=[], graph=None, **kwargs):
    '''
    Make a scatter plot of the Data object

    xitvl: 2 component list [a,b] or tuple (a,b)
        restriction to the x-axis (a < x < b)

    yitvl: 2 component list [a,b] or tuple (a,b)
        restriction to the y-axis (a < y < b)

    graph: a Axes object in Matplotlib
    '''
    data = snip(data, xitvl, yitvl)

    if graph: return graph.scatter(data.xval, data.yval, **kwargs)
    else: plt.scatter(data.xval, data.yval, **kwargs)

def errorbar(data, xitvl=[], yitvl=[], graph=None, **kwargs):
    '''
    Return a Line2D object in Matplotlib or plot the Data object (graph = None) with errorbar

    xitvl: 2 component list [a,b] or tuple (a,b)
        restriction to the x-axis (a < x < b)

    yitvl: 2 component list [a,b] or tuple (a,b)
        restriction to the y-axis (a < y < b)

    graph: a Axes object in Matplotlib
    '''
    data = snip(data, xitvl, yitvl)

    if graph: return graph.errorbar(data.xval, data.yval, xerr=data.xerr, yerr=data.yerr, **kwargs)
    else: plt.errorbar(data.xval, data.yval, xerr=data.xerr, yerr=data.yerr, **kwargs)

def polyplot(coeffs, itvl, graph=None, point=100, **kwargs):
    '''
    Return a Line2D object in Matplotlib of a polynomial or plot the polynomial (graph = None)

    coeffs: 1D array
        the coefficients of the polynomial from highest order to lowest order

    itvl: 2 component list [a,b] or tuple (a,b)
        the range of the x-axis

    graph: a Axes object in Matplotlib

    point: int
        the number of points to plot
    '''
    assert np.shape(itvl)==(2,), "'itvl' error"
    temp_x = np.linspace(itvl[0], itvl[1], point)
    temp_y = (np.poly1d(coeffs))(temp_x)

    if graph: return graph.plot(temp_x, temp_y, **kwargs)
    else: plt.plot(temp_x, temp_y, **kwargs)

def funcplot(func, itvl, param={}, graph=None, point=100, **kwargs):
    '''
    Return a Line2D object in Matplotlib of a function or plot the function (graph = None)

    itvl: 2 component list [a,b] or tuple (a,b)
        the range of the x-axis

    param: a dictionary
        the value of parameters in the function

    graph: a Axes object in Matplotlib

    point: int
        the number of points to plot
    '''
    assert np.shape(itvl)==(2,), "'itvl' error"
    temp_x = np.linspace(itvl[0], itvl[1], point)
    temp_y = func(temp_x, **param)

    if graph: return graph.plot(temp_x, temp_y, **kwargs)
    else: plt.plot(temp_x, temp_y, **kwargs)

def directplot(fitting_result, itvl, graph=None, point=100, **kwargs):
    '''
    Return a Line2D object in Matplotlib from a FitResult object or plot it (graph = None)

    itvl: 2 component list [a,b] or tuple (a,b)
        the range of the x-axis

    graph: a Axes object in Matplotlib

    point: int
        the number of points to plot
    '''
    if int(fitting_result)==2:
        func = lambda x: (fitting_result.funcinfo)(x, *(fitting_result.coeffs()))
        funcplot(func, itvl, {}, graph, point, **kwargs)
    else:
        polyplot(fitting_result.coeffs(), itvl, graph, point, **kwargs)

def multiplot(*data, xitvl=[], yitvl=[], graph=None, err=False, **kwargs):
    '''
    Return a Line2D object in Matplotlib or plot the Data objects (graph = None)

    xitvl: 2 component list [a,b] or tuple (a,b)
        restriction to the x-axis (a < x < b)

    yitvl: 2 component list [a,b] or tuple (a,b)
        restriction to the y-axis (a < y < b)
    
    graph: a Axes object in Matplotlib

    err: Boolean
        whether to plot the errorbar or not
    '''
    for single_data in data:
        if err: errorbar(single_data, xitvl, yitvl, graph, **kwargs)
        else: plot(single_data, xitvl, yitvl, graph, **kwargs)

def show():
    '''
    Equivalent to plt.show()
    '''
    plt.show()


## Smoothing ##

def smooth(data, weight=[1., 1., 1., 1., 1.]):
    '''
    Return a smoothed Data object with respect to a given weight

    weight: a list
        the weight of adjacent data points
    '''
    data = sort(data)
    weight = np.array(weight)/np.sum(weight)
    val, errsq = (array(data))[:,:2], ((array(data))[:,2:])**2
    new_val = _weight(val, weight)
    new_errsq = _weight(errsq, weight**2)
    
    return Data(new_val[:,0], new_val[:,1], np.sqrt(new_errsq[:,0]), np.sqrt(new_errsq[:,1]))

def savgol(data, window, order, deriv=0):
    '''
    Return a Data object after applying the Savitzky-Golay method
    
    window: int (odd number)
        the number of sample points
    
    order: int
        the order of fitting polynomial
    
    deriv: 0, 1, or 2 (<= order)
        the order of derivation to compute
    '''
    assert window%2==1, "'window' error"
    data = sort(data)
    val, errsq = (array(data))[:,:2], ((array(data))[:,2:])**2
    if deriv==0:
        new_val = (spsgn.savgol_filter(val, window, order, axis=0, mode='nearest'))[window//2:len(data)-window//2]
        new_errsq = _weight(errsq, (spsgn.savgol_coeffs(window, order))**2)
    elif deriv==1:
        new_val, new_errsq = np.zeros((len(data)-window+1, 2)), np.zeros((len(data)-window+1, 2))
        
        new_val[:,0] = (spsgn.savgol_filter(val[:,0], window, order, mode='nearest'))[window//2:len(data)-window//2]
        dval = (spsgn.savgol_filter(val, window, order, 1, axis=0, mode='nearest'))[window//2:len(data)-window//2]
        new_val[:,1] = dval[:,1]/dval[:,0]
        
        new_errsq[:,0] = _weight(errsq[:,0], (spsgn.savgol_coeffs(window, order))**2)
        derrsq = _weight(errsq, (spsgn.savgol_coeffs(window, order, 1))**2)
        new_errsq[:,1] = derrsq[:,1]/(dval[:,0])**2+derrsq[:,0]*(dval[:,1]/(dval[:,0])**2)**2
    elif deriv==2:
        new_val, new_errsq = np.zeros((len(data)-window+1, 2)), np.zeros((len(data)-window+1, 2))

        new_val[:,0] = (spsgn.savgol_filter(val[:,0], window, order, mode='nearest'))[window//2:len(data)-window//2]
        dval = (spsgn.savgol_filter(val, window, order, 1, axis=0, mode='nearest'))[window//2:len(data)-window//2]
        ddval = (spsgn.savgol_filter(val, window, order, 2, axis=0, mode='nearest'))[window//2:len(data)-window//2]
        new_val[:,1] = ddval[:,1]/(dval[:,0])**2-ddval[:,0]*dval[:,1]/(dval[:,0])**3

        new_errsq[:,0] = _weight(errsq[:,0], (spsgn.savgol_coeffs(window, order))**2)
        derrsq = _weight(errsq, (spsgn.savgol_coeffs(window, order, 1))**2)
        dderrsq = _weight(errsq, (spsgn.savgol_coeffs(window, order, 2))**2)
        new_errsq[:,1] = dderrsq[:,1]/(dval[:,0])**4+dderrsq[:,0]*(dval[:,1]/(dval[:,0])**3)**2+derrsq[:,1]*(ddval[:,0]/(dval[:,0])**3)**2+derrsq[:,0]*(2.*ddval[:,1]/(dval[:,0])**3-3.*ddval[:,0]*dval[:,1]/(dval[:,0])**4)**2
    else: assert False, "'deriv' error"

    return Data(new_val[:,0], new_val[:,1], np.sqrt(new_errsq[:,0]), np.sqrt(new_errsq[:,1]))

def nbrsmooth(data, radius):
    '''
    Return a Data object after taking moving average with fixed radius

    radius: float
        the radius of sample space
    '''
    data = sort(data)
    x_val, y_val, y_errsq = (array(data))[:,0], (array(data))[:,1], ((array(data))[:,3])**2
    new_yval, new_yerrsq = np.zeros(len(data)), np.zeros(len(data))
    bound = _self_neighbor(x_val, radius)
    for n in range(len(data)):
        lower, upper = (bound[n])[0], (bound[n])[1]
        new_yval[n] = np.mean(y_val[lower:upper+1])
        new_yerrsq[n] = (np.var(y_val[lower:upper+1])+np.sum(y_errsq[lower:upper+1]))/(upper-lower+1)

    return Data(x_val, new_yval, data.xerr, np.sqrt(new_yerrsq))

def polysmooth(data, radius, order, deriv=0):
    '''
    Return a Data object after taking polynomial regression with fixed radius
    
    radius: float
        the radius of sample space

    order: int
        the order of fitting polynomial

    deriv: int (<= order)
        the order of derivation to compute
    '''
    assert deriv<=order, "'deriv' bigger than 'order'"
    data = sort(data)
    x_val, y_val, y_errsq = (array(data))[:,0], (array(data))[:,1], ((array(data))[:,3])**2
    new_yval, new_yerrsq = np.zeros(len(data)), np.zeros(len(data))
    bound = _self_neighbor(x_val, radius)
    for n in range(len(data)):
        lower, upper = (bound[n])[0], (bound[n])[1]
        coeffs, cov = np.polyfit(x_val[lower:upper+1]-x_val[n], y_val[lower:upper+1], order, w=1./y_errsq[lower:upper+1], cov=True)
        new_yval[n] = np.math.factorial(deriv)*coeffs[order-deriv]
        new_yerrsq[n] = cov[order-deriv, order-deriv]+np.mean(y_errsq[lower:upper+1])

    return Data(x_val, new_yval, data.xerr, np.sqrt(new_yerrsq))


## Interpolation ##

def interpolate(data, xval, xerr=0., window=2, order=1, extrap=False):
    '''
    Return y_val and y_err with interpolation using adjacent points

    xval: float or a list

    xerr: float or a list    

    window: int (even number)
        the number of sample points

    order: int (0 < order < window-1 except for window=2, order=1)
        the order of fitting polynomial

    extrap: Boolean
        whether accept extrapolation or not
    '''
    if isinstance(xval, float): Xval = np.array([xval])
    else: Xval = np.array(xval)
    if isinstance(xerr, float): Xerr = np.full(len(Xval), xerr)
    else:
        assert len(Xval)==len(xerr), "'xval' or 'xerr' error"
        Xerr = np.array(xerr)
    
    if not ((window==2) and (order==1)): assert (window%2==0) and (order<window-1) and window<len(data)//2, "'window' error"
    data = array(sort(data))
    x_val, y_val = data[:,0], data[:,1]
    if not extrap: assert (np.amax(Xval)<np.amax(x_val)) and (np.amin(Xval)>np.amin(x_val)), "'extrap' error"

    indices = _adjacent(Xval, x_val, window)
    if (window==2) and (order==1):
        x_err, y_err = data[:,2], data[:,3]
        Yval, Yerr = np.zeros(len(Xval)), np.zeros(len(Xval))
        for n in range(len(Xval)):
            ind = indices[n]
            a, b, c = u.ufloat(x_val[ind], x_err[ind]), u.ufloat(x_val[ind+1], x_err[ind+1]), u.ufloat(Xval[n], Xerr[n])
            fa, fb = u.ufloat(y_val[ind], y_err[ind]), u.ufloat(y_val[ind+1], y_err[ind+1])
            fc = ((c-a)*fb+(b-c)*fa)/(b-a)
            Yval[n], Yerr[n] = fc.n, fc.s
        
        if isinstance(xval, float): return Yval[0], Yerr[0]
        else: return (np.dstack(Yval, Yerr))[0]
    else:
        Xerrsq, x_errsq, y_errsq = Xerr**2, (data[:,2])**2, (data[:,3])**2
        Yval, Yerrsq = np.zeros(len(Xval)), np.zeros(len(Xval))
        for n in range(len(Xval)):
            ind = indices[n]
            coeffs, cov = np.polyfit(x_val[ind:ind+window]-Xval[n], y_val[ind:ind+window], order, w=1./y_errsq[ind:ind+window], cov=True)
            Yval[n] = coeffs[-1]
            Yerrsq[n] = coeffs[-2]**2*(Xerrsq[n]+np.mean(x_errsq[ind:ind+window]))+np.mean(y_errsq[ind:ind+window])+cov[-1,-1]

        if isinstance(xval, float): return Yval[0], np.sqrt(Yerrsq[0])
        else: return (np.dstack(Yval, np.sqrt(Yerrsq)))[0]

def xyzinterp(xydata, yzdata, window=2, order=1, extrap=False):
    '''
    Return the Data object of xzdata from xydata(smaller) and yzdata(larger) with interpolation using adjacent points

    window: int (even number)
        the number of sample points

    order: int (0 < order < window-1 except for window=2, order=1)
        the order of fitting polynomial

    extrap: Boolean
        whether accept extrapolation or not
    '''
    if not ((window==2) and (order==1)): assert (window%2==0) and (order<window-1) and window<len(yzdata)//2, "'window' error"
    xydata, yzdata = array(sort(xydata, False)), array(sort(yzdata))
    y1_val, y2_val, z_val = xydata[:,1], yzdata[:,0], yzdata[:,1]
    if not extrap: assert (np.amax(y1_val)<np.amax(y2_val)) and (np.amin(y1_val)>np.amin(y2_val)), "'extrap' error"

    indices = _adjacent(y1_val, y2_val, window)
    if (window==2) and (order==1):
        y1_err, y2_err, z_err = xydata[:,3], yzdata[:,2], yzdata[:,3]
        new_yval, new_yerr = np.zeros(len(xydata)), np.zeros(len(xydata))
        for n in range(len(xydata)):
            ind = indices[n]
            a, b, c = u.ufloat(y2_val[ind], y2_err[ind]), u.ufloat(y2_val[ind+1], y2_err[ind+1]), u.ufloat(y1_val[n], y1_err[n])
            fa, fb = u.ufloat(z_val[ind], z_err[ind]), u.ufloat(z_val[ind+1], z_err[ind+1])
            fc = ((c-a)*fb+(b-c)*fa)/(b-a)
            new_yval[n], new_yerr[n] = fc.n, fc.s
        
        return Data(xydata[:,0], new_yval, xydata[:,2], new_yerr)
    else:
        y1_errsq, y2_errsq, z_errsq = (xydata[:,3])**2, (yzdata[:,2])**2, (yzdata[:,3])**2
        new_yval, new_yerrsq = np.zeros(len(xydata)), np.zeros(len(xydata))
        for n in range(len(xydata)):
            ind = indices[n]
            coeffs, cov = np.polyfit(y2_val[ind:ind+window]-y1_val[n], z_val[ind:ind+window], order, w=1./z_errsq[ind:ind+window], cov=True)
            new_yval[n] = coeffs[-1]
            new_yerrsq[n] = coeffs[-2]**2*(y1_errsq[n]+np.mean(y2_errsq[ind:ind+window]))+np.mean(z_errsq[ind:ind+window])+cov[-1,-1]

        return Data(xydata[:,0], new_yval, xydata[:,2], np.sqrt(new_yerrsq))

def wxyinterp(xydata, wxdata, window=2, order=1, extrap=False):
    '''
    Return the Data object of wydata from xydata(smaller) and wxdata(larger) with interpolation using adjacent points

    window: int (even number)
        the number of sample points

    order: int (0 < order < window-1 except for window=2, order=1)
        the order of fitting polynomial

    extrap: Boolean
        whether accept extrapolation or not
    '''
    if not ((window==2) and (order==1)): assert (window%2==0) and (order<window-1) and window<len(wxdata)//2, "'window' error"
    xydata, wxdata = array(sort(xydata, False)), array(switch(sort(switch(wxdata))))
    x1_val, x2_val, w_val = xydata[:,0], wxdata[:,1], wxdata[:,0]
    if not extrap: assert (np.amax(x1_val)<np.amax(x2_val)) and (np.amin(x1_val)>np.amin(x2_val)), "'extrap' error"

    indices = _adjacent(x1_val, x2_val, window)
    if (window==2) and (order==1):
        x1_err, x2_err, w_err = xydata[:,2], wxdata[:,3], wxdata[:,2]
        new_xval, new_xerr = np.zeros(len(xydata)), np.zeros(len(xydata))
        for n in range(len(xydata)):
            ind = indices[n]
            a, b, c = u.ufloat(x2_val[ind], x2_err[ind]), u.ufloat(x2_val[ind+1], x2_err[ind+1]), u.ufloat(x1_val[n], x1_err[n])
            fa, fb = u.ufloat(w_val[ind], w_err[ind]), u.ufloat(w_val[ind+1], w_err[ind+1])
            fc = ((c-a)*fb+(b-c)*fa)/(b-a)
            new_xval[n], new_xerr[n] = fc.n, fc.s
        
        return Data(new_xval, xydata[:,1], new_xerr, xydata[:,3])
    else:
        x1_errsq, x2_errsq, w_errsq = (xydata[:,2])**2, (wxdata[:,3])**2, (wxdata[:,2])**2
        new_xval, new_xerrsq = np.zeros(len(xydata)), np.zeros(len(xydata))
        for n in range(len(xydata)):
            ind = indices[n]
            coeffs, cov = np.polyfit(x2_val[ind:ind+window]-x1_val[n], w_val[ind:ind+window], order, w=1./w_errsq[ind:ind+window], cov=True)
            new_xval[n] = coeffs[-1]
            new_xerrsq[n] = coeffs[-2]**2*(x1_errsq[n]+np.mean(x2_errsq[ind:ind+window]))+np.mean(w_errsq[ind:ind+window])+cov[-1,-1]

        return Data(new_xval, xydata[:,1], np.sqrt(new_xerrsq), xydata[:,3])

def nbrinterp(data, radius, xval, xerr=0., order=2):
    '''
    Return the Data object of xzdata from xydata(smaller) and yzdata(larger) with interpolation inside fixed radius

    radius: float
        the radius of sample space

    xval: float or a list

    xerr: float or a list

    order: int (nonzero number)
        the order of fitting polynomial
    '''
    if isinstance(xval, float): Xval = np.array([xval])
    else: Xval = np.array(xval)
    if isinstance(xerr, float): Xerr = np.full(len(Xval), xerr)
    else:
        assert len(Xval)==len(xerr), "'xval' or 'xerr' error"
        Xerr = np.array(xerr)

    assert order>0, "'order' error"
    data = array(sort(data))
    x_val, y_val, x_errsq, y_errsq = data[:,0], data[:,1], (data[:,2])**2, (data[:,3])**2
    assert (np.amax(Xval)<np.amax(x_val)) and (np.amin(Xval)>np.amin(x_val)), "only interpolation allowed"
    
    bound = _neighbor(Xval, x_val, radius)
    Xerrsq, Yval, Yerrsq = Xerr**2, np.zeros(len(Xval)), np.zeros(len(Xval))
    for n in range(len(Xval)):
        lower, upper = (bound[n])[0], (bound[n])[1]
        assert upper-lower>2*order, "'radius' error"
        coeffs, cov = np.polyfit(x_val[lower:upper+1]-Xval[n], y_val[lower:upper+1], order, w=1./y_errsq[lower:upper+1], cov=True)
        Yval[n] = coeffs[-1]
        Yerrsq[n] = coeffs[-2]**2*(Xerrsq[n]+np.mean(x_errsq[lower:upper+1]))+np.mean(y_errsq[lower:upper+1])+cov[-1,-1]

    if isinstance(xval, float): return Yval[0], np.sqrt(Yerrsq[0])
    else: return (np.dstack(Yval, np.sqrt(Yerrsq)))[0]

def nbrxyzinterp(xydata, yzdata, radius, order=2):
    '''
    Return the Data object of xzdata from xydata(smaller) and yzdata(larger) with interpolation inside fixed radius

    radius: float
        the radius of sample space

    order: int (nonzero number)
        the order of fitting polynomial
    '''
    assert order>0, "'order' error"
    xydata, yzdata = array(sort(xydata, False)), array(sort(yzdata))
    y1_val, y2_val, z_val = (array(xydata))[:,1], (array(yzdata))[:,0], (array(yzdata))[:,1]
    y1_errsq, y2_errsq, z_errsq = ((array(xydata))[:,3])**2, ((array(yzdata))[:,2])**2, ((array(yzdata))[:,3])**2
    new_yval, new_yerrsq = np.zeros(len(xydata)), np.zeros(len(xydata))
    assert (np.amax(y1_val)<np.amax(y2_val)) and (np.amin(y1_val)>np.amin(y2_val)), "only interpolation allowed"

    bound = _neighbor(y1_val, y2_val, radius)
    for n in range(len(xydata)):
        lower, upper = (bound[n])[0], (bound[n])[1]
        assert upper-lower>2*order, "'radius' error"
        coeffs, cov = np.polyfit(y2_val[lower:upper+1]-y1_val[n], z_val[lower:upper+1], order, w=1./z_errsq[lower:upper+1], cov=True)
        new_yval[n] = coeffs[-1]
        new_yerrsq[n] = coeffs[-2]**2*(y1_errsq[n]+np.mean(y2_errsq[lower:upper+1]))+np.mean(z_errsq[lower:upper+1])+cov[-1,-1]

    return Data(xydata[:,0], new_yval, xydata[:,2], np.sqrt(new_yerrsq))

def nbrwxyinterp(xydata, wxdata, radius, order=2):
    '''
    Return the Data object of wydata from xydata(smaller) and wxdata(larger) with interpolation inside fixed radius

    radius: float
        the radius of sample space

    order: int (nonzero number)
        the order of fitting polynomial
    '''
    assert order>0, "'order' error"
    xydata, wxdata = array(sort(xydata, False)), array(switch(sort(switch(wxdata))))
    x1_val, x2_val, w_val = xydata[:,0], wxdata[:,1], wxdata[:,0]
    x1_errsq, x2_errsq, w_errsq = (xydata[:,2])**2, (wxdata[:,3])**2, (wxdata[:,2])**2
    new_xval, new_xerrsq = np.zeros(len(xydata)), np.zeros(len(xydata))
    assert (np.amax(x1_val)<np.amax(x2_val)) and (np.amin(x1_val)>np.amin(x2_val)), "only interpolation allowed"

    bound = _neighbor(x1_val, x2_val, radius)
    for n in range(len(xydata)):
        lower, upper = (bound[n])[0], (bound[n])[1]
        assert upper-lower>2*order, "'radius' error"
        coeffs, cov = np.polyfit(x2_val[lower:upper+1]-x1_val[n], w_val[lower:upper+1], order, w=1./w_errsq[lower:upper+1], cov=True)
        new_xval[n] = coeffs[-1]
        new_xerrsq[n] = coeffs[-2]**2*(x1_errsq[n]+np.mean(x2_errsq[lower:upper+1]))+np.mean(w_errsq[lower:upper+1])+cov[-1,-1]

    return Data(new_xval, xydata[:,1], np.sqrt(new_xerrsq), xydata[:,3])


## Fitting ##

def fitting(data, func, threshold=5., **params):
    '''
    Return a FitResult object that minimizes the least-squared cost function of the given function

    func: y = f(x, *params)
        the function used to fit the data

    threshold: float
        the maximum acceptable value of chisq/dof

    **params: param1 = init_guess1, param2 = init_guess2, ....
        initial guess of the parameters
    '''
    ave_chisq = 100000.
    count = 0
    while ave_chisq>threshold:
        if count>0:
            for param in params.keys():
                params[param] *= 1+0.01*(np.random.rand()-0.5)
        temp_params = (spopt.curve_fit(func, data.xval, data.yval, list(params.values()), data.yerr, True))[0]
        ave_chisq = _chisq(data, (data.yerr)**2, func, temp_params)/(len(data)-len(params))
        count += 1
        assert count<10, "'params' or 'threshold' error"
    
    temp_func = lambda x: func(x, *temp_params)
    slope = spmisc.derivative(temp_func, data.xval, 1E-6)
    errsq = slope**2*(data.xerr)**2+(data.yerr)**2
    
    ave_chisq = 100000.
    count = 0
    while ave_chisq>threshold:
        if count>0:
            for n in range(len(temp_params)):
                temp_params[n] *= 1+0.001*(np.random.rand()-0.5)
        temp_params, cov = spopt.curve_fit(func, data.xval, data.yval, temp_params, np.sqrt(errsq), True)
        ave_chisq = _chisq(data, (data.yerr)**2, func, temp_params)/(len(data)-len(params))
        count += 1
        assert count<20, "'params' or 'threshold' error"

    count = 0
    for param in params.keys():
        params[param] = np.array([temp_params[count], np.sqrt(cov[count, count])])
        count += 1

    return FitResult(func, params, ave_chisq)

def polyfit(data, order):
    '''
    Return a FitResult object that minimizes the least-squared cost function of a polynomial
    
    order: int (nonzero number)
        the order of fitting polynomial
    '''
    coeffs = np.polyfit(data.xval, data.yval, order, w=1./data.yerr)
    poly = np.poly1d(coeffs)
    errsq = ((poly.deriv())(data.xval))**2*(data.xerr)**2+(data.yerr)**2
    coeffs, cov = np.polyfit(data.xval, data.yval, order, w=1./np.sqrt(errsq), cov=True)
    ave_chisq = _chisq(data, errsq, poly)/(len(data)-order-1)
    param = {}
    for n in range(order+1):
        param[order-n] = np.array([coeffs[n], np.sqrt(cov[n,n])])
    
    return FitResult('poly', param, ave_chisq)

def linereg(data):
    '''
    Return a FitResult object corresponding to the linear regression method
    '''
    mval, bval, r = (spstat.linregress(data.xval, data.yval))[:3]
    errsq = mval**2*(data.xerr)**2+(data.yerr)**2
    temp = len(data)*np.sum((data.xval)**2)-(np.sum(data.xval))**2
    merr, berr = np.sqrt(np.sum(errsq)/temp), np.sqrt(np.sum((data.xval)**2)*np.mean(errsq)/temp)
    func = lambda x: mval*x+bval
    ave_chisq = _chisq(data, errsq, func)/(len(data)-2)
    param = {'m': np.array([mval, merr]), 'b': np.array([bval, berr])}

    return FitResult('line', param, ave_chisq, r**2)
    

